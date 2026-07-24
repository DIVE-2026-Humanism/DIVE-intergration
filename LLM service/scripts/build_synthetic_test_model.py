#!/usr/bin/env python3
"""KCB 컬럼 정의서 기반 테스트 전용 합성데이터와 분류 아티팩트를 만든다.

이 스크립트의 출력은 연결·학습 파이프라인 검증용이며 실제 성능이나 정책 판단에
사용할 수 없다. 대회 당일에는 공식 KCB 파일로 기본 학습 명령을 다시 실행한다.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
import unicodedata
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.external import validate_external  # noqa: E402
from src.infer import infer_pipeline  # noqa: E402
from src.io_load import (  # noqa: E402
    load_config,
    load_main,
    normalized_glob,
    write_data_profile,
    write_json,
)
from src.labeling import CLASS_ORDER  # noqa: E402
from src.leakage import leakage_report  # noqa: E402
from src.preprocess import preprocess_main  # noqa: E402
from src.train import prepare_full, train_pipeline  # noqa: E402
from src.validate import validate_pipeline  # noqa: E402


WARNING = "SYNTHETIC_TEST_ONLY: 실제 성능·경제상태·정책 판단에 사용 금지"
DEFAULT_OUTPUT = ROOT / "Dataset" / "KCB_테스트_합성데이터.csv"
BUSAN_GU_CODES = np.asarray(
    [26110, 26140, 26170, 26200, 26230, 26260, 26290, 26320, 26350, 26380, 26410, 26440, 26470, 26500, 26530, 26710],
    dtype=int,
)


def _nfc(value: Any) -> str:
    return unicodedata.normalize("NFC", str(value).strip())


def locate_dictionary() -> Path:
    candidates = normalized_glob(ROOT / "Dataset", "*KCB*합성데이터*컬럼*자료.xlsx")
    if len(candidates) != 1:
        raise FileNotFoundError(f"KCB 컬럼 정의서가 정확히 하나 필요합니다: {[path.name for path in candidates]}")
    return candidates[0]


def dictionary_columns(path: Path) -> list[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = None
    column_index = None
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        normalized = [_nfc(value) if value is not None else "" for value in row]
        if "컬럼코드" in normalized:
            header_row = row_number
            column_index = normalized.index("컬럼코드")
            break
    if header_row is None or column_index is None:
        raise ValueError("컬럼 정의서에서 '컬럼코드' 헤더를 찾지 못했습니다.")
    columns: list[str] = []
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        value = row[column_index]
        if value is not None and str(value).strip():
            columns.append(_nfc(value))
    if len(columns) != 46 or len(columns) != len(set(columns)):
        raise ValueError(f"샘플 정의서에서 중복 없는 46개 컬럼을 기대했지만 {len(columns)}개입니다.")
    return columns


def _bounded_income(target: str, ext_inc: float, q1: float, q5: float, rng: np.random.Generator) -> float:
    if target == "V1":
        return rng.uniform(q1 * 0.75, ext_inc * 0.92)
    if target == "V2":
        return rng.uniform(max(q1 * 1.15, ext_inc * 0.90), min(q5 * 0.85, ext_inc * 1.15))
    if target == "V3":
        return rng.uniform(q1 * 0.55, q1 * 0.90)
    if target == "S1":
        return rng.uniform(q5 * 1.08, q5 * 1.35)
    if target == "S2":
        return rng.uniform(max(q1 * 1.20, ext_inc * 0.85), min(q5 * 0.82, ext_inc * 1.12))
    if target == "S3":
        return rng.uniform(ext_inc * 1.05, min(q5 * 0.82, ext_inc * 1.25))
    raise ValueError(f"지원하지 않는 합성 목표 유형: {target}")


def _allocate_debt(total: int, target: str) -> tuple[int, int, int]:
    if total <= 0:
        return 0, 0, 0
    if target in {"V1", "V2"}:
        credit = int(round(total * 0.72))
        policy = total - credit
        return credit, 0, policy
    credit = int(round(total * 0.45))
    mortgage = total - credit
    return credit, mortgage, 0


def synthetic_row(
    target: str,
    row_number: int,
    columns: list[str],
    config: dict[str, Any],
    rng: np.random.Generator,
) -> dict[str, Any]:
    row: dict[str, Any] = {column: 0 for column in columns}
    age = int(rng.choice(config["filters"]["youth_age_codes"]))
    scale = float(config["external"]["amount_scale_to_kcb"])
    anchor = config["external"]["anchors"][age]
    ext_inc = float(anchor["inc"]) * scale
    ext_q1 = float(anchor["q1"]) * scale
    ext_q5 = float(anchor["q5"]) * scale
    ext_card = float(anchor["card"]) * scale
    ext_debt = float(config["external"]["debt"][age]) * scale

    income = max(1_200, int(round(_bounded_income(target, ext_inc, ext_q1, ext_q5, rng))))
    change_ranges = {
        "V1": (-0.30, -0.05),
        "V2": (-0.08, 0.08),
        "V3": (-0.30, -0.04),
        "S1": (0.05, 0.24),
        "S2": (0.03, 0.18),
        "S3": (0.01, 0.16),
    }
    income_change = float(rng.uniform(*change_ranges[target]))
    previous_income = max(1_000, int(round(income / (1.0 + income_change))))

    dsr_ranges = {
        "V1": (0.25, 0.48),
        "V2": (0.88, 0.96),
        "V3": (0.00, 0.00),
        "S1": (0.00, 0.00),
        "S2": (0.04, 0.14),
        "S3": (0.00, 0.00),
    }
    dsr = float(rng.uniform(*dsr_ranges[target]))
    debt_ranges = {
        "V1": (0.90, 1.50),
        "V2": (1.55, 2.40),
        "V3": (0.00, 0.00),
        "S1": (0.00, 0.00),
        "S2": (0.20, 0.70),
        "S3": (0.00, 0.00),
    }
    debt_total = int(round(ext_debt * rng.uniform(*debt_ranges[target])))
    loan_count = int(rng.integers(2, 6)) if target in {"V1", "V2"} else int(rng.integers(1, 4)) if target == "S2" else 0
    credit_balance, mortgage_balance, policy_balance = _allocate_debt(debt_total, target)

    card_ratio_ranges = {
        "V1": (0.55, 0.82),
        "V2": (0.35, 0.62),
        "V3": (0.78, 1.05),
        "S1": (0.20, 0.38),
        "S2": (0.24, 0.48),
        "S3": (0.20, 0.38),
    }
    card_total = int(round(income * rng.uniform(*card_ratio_ranges[target])))
    if target == "V3":
        card_total = max(card_total, int(round(ext_card * 1.12)))
    if target == "S3":
        card_total = min(card_total, int(round(ext_card * 0.82)))
    credit_card = int(round(card_total * rng.uniform(0.68, 0.82)))
    check_card = card_total - credit_card
    cash_service = int(round(income * rng.uniform(0.03, 0.08))) if target == "V1" else 0

    score_ranges = {
        "V1": (430, 565),
        "V2": (500, 625),
        "V3": (585, 685),
        "S1": (805, 920),
        "S2": (735, 830),
        "S3": (760, 860),
    }
    credit_score = int(round(rng.uniform(*score_ranges[target])))

    if target == "V1":
        loan_delq = int(rng.integers(2, 5))
        card_delq = int(rng.integers(0, 3))
        delq_days = int(rng.integers(75, 220))
        delq_total = int(round((income / 12.0) * rng.uniform(1.1, 3.0)))
    else:
        loan_delq = card_delq = delq_days = delq_total = 0
    loan_delq_amount = int(round(delq_total * 0.75))
    card_delq_amount = delq_total - loan_delq_amount

    owner = target == "S1"
    house_value = int(rng.integers(250_000, 800_001)) if owner else 0
    net_house = int(round(house_value * rng.uniform(0.72, 0.98))) if owner else 0
    residence_code = int(rng.choice(BUSAN_GU_CODES))
    work_code = int(rng.choice(BUSAN_GU_CODES))
    household_size = 1 if rng.random() < 0.70 else int(rng.integers(2, 5))
    agreement_multiplier = float(rng.uniform(1.03, 1.20))
    repayment = int(round(income * dsr))
    installment = int(round(credit_card * rng.uniform(0.12, 0.30)))

    row.update(
        {
            "성별": int(rng.integers(1, 3)),
            "연령대": age,
            "직업군": int(rng.choice([910, 920, 930, 940, 950])),
            "직업군상세": int(rng.choice([911, 921, 931, 941, 951])),
            "거주지 시군구 코드": residence_code,
            "거주지행정동": residence_code * 1000 + int(rng.integers(101, 999)),
            "근무지 시군구 코드": work_code,
            "근무지 행정동": work_code * 1000 + int(rng.integers(101, 999)),
            "추정가구원수": household_size,
            "추정월소득": int(round(income / 12.0)),
            "증빙연소득": int(round(income * rng.uniform(0.75, 1.02))),
            "추정 연소득": income,
            "2년전 추정 연소득 금액": previous_income,
            "총자산평가금액(주택)": house_value,
            "순자산평가금액(주택)": net_house,
            "자가거주여부": 1 if owner else 3,
            "현 거주지의 아파트여부": 1 if owner or rng.random() < 0.55 else 0,
            "현 거주지의 매매가(국토부 실거래가) 또는 공시가격": house_value,
            "차량보유(국산/수입)": int(rng.choice([0, 0, 0, 1, 2])),
            "추정 LTV": round(100.0 * mortgage_balance / house_value, 2) if house_value else 0,
            "추정DTI": round(dsr * 100.0, 2),
            "신용평점": credit_score,
            "총대출건수": loan_count,
            "신용대출-총대출약정액": int(round(credit_balance * agreement_multiplier)),
            "신용대출-총대출잔액": credit_balance,
            "주택담보대출-총대출약정액": int(round(mortgage_balance * agreement_multiplier)),
            "주택담보대출-총대출잔액": mortgage_balance,
            "정책자금대출-총대출약정액": int(round(policy_balance * agreement_multiplier)),
            "정책자금대출-총대출잔액": policy_balance,
            "총 대출 상환금액 (최근 12개월)": repayment,
            "최근 12개월 신용카드소비금액": credit_card,
            "최근 12개월 체크카드소비금액": check_card,
            "최근 12개월 일시불이용금액": credit_card - installment,
            "최근 12개월 할부이용금액": installment,
            "최근 12개월 현금서비스이용금액": cash_service,
            "대출연체건수": loan_delq,
            "카드연체건수": card_delq,
            "연체일수": delq_days,
            "대출연체금액": loan_delq_amount,
            "카드연체금액": card_delq_amount,
            "Thin Filer 여부": 0,
            "파산, 개인회생 신청 여부": 1 if target == "V1" and rng.random() < 0.12 else 0,
            "2년내 현거주지평균실거래가": int(round(house_value * 0.96)) if owner else -99_999_999,
            "2년내 현거주지평균전세거래가": int(round(house_value * 0.62)) if owner else -99_999_999,
            "2년내 직장명이력건수": int(rng.integers(0, 5)),
            "2년내 이직후 소득 증감액": income - previous_income,
        }
    )
    row["합성목표유형"] = target
    row["합성행ID"] = f"SYN-{target}-{row_number:05d}"
    row["데이터구분"] = "SYNTHETIC_TEST_ONLY"
    return row


def generate_frame(rows_per_type: int, seed: int, columns: list[str], config: dict[str, Any]) -> pd.DataFrame:
    if rows_per_type < 30:
        raise ValueError("교차검증을 위해 유형별 최소 30행이 필요합니다.")
    rng = np.random.default_rng(seed)
    rows = [
        synthetic_row(target, number, columns, config, rng)
        for target in CLASS_ORDER
        for number in range(1, rows_per_type + 1)
    ]
    frame = pd.DataFrame(rows, columns=columns + ["합성목표유형", "합성행ID", "데이터구분"])
    return frame.sample(frac=1.0, random_state=seed).reset_index(drop=True)


def validate_generated(frame: pd.DataFrame, config: dict[str, Any]) -> dict[str, Any]:
    processed, stats = preprocess_main(frame, config)
    _, _, _, labels = prepare_full(processed, config)
    table = pd.crosstab(processed["합성목표유형"], labels["LABEL"].fillna("HOLDOUT"))
    counts = labels["LABEL"].value_counts().reindex(CLASS_ORDER, fill_value=0)
    minimum = max(10, len(frame) // 120)
    insufficient = {code: int(counts[code]) for code in CLASS_ORDER if int(counts[code]) < minimum}
    if insufficient:
        raise ValueError(f"합성 레이블 표본이 부족합니다: {insufficient}\n{table.to_string()}")
    return {
        "preprocess": stats,
        "rule_label_counts": {code: int(counts[code]) for code in CLASS_ORDER},
        "target_vs_rule": {str(index): {str(key): int(value) for key, value in row.items()} for index, row in table.to_dict("index").items()},
        "holdout_rows": int(labels["LABEL"].isna().sum()),
    }


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_artifacts(config: dict[str, Any], output: Path, iterations: int) -> dict[str, Any]:
    config["project"]["data_kind"] = "synthetic_test"
    config["paths"]["main_csv"] = output.name
    config["training"]["iterations"] = int(iterations)
    config["training"]["early_stopping_rounds"] = max(20, min(60, iterations // 4))
    validate_external(config)
    write_data_profile(config)
    raw, metadata = load_main(config)
    processed, preprocess_stats = preprocess_main(raw, config)
    training_result = train_pipeline(processed, config)
    validation_path = validate_pipeline(processed, config)
    leakage_path = leakage_report(processed, config)
    predictions_path = infer_pipeline(processed, config)
    return {
        "input": metadata,
        "preprocess": preprocess_stats,
        "metrics": training_result["metrics"],
        "validation_report": str(validation_path),
        "leakage_report": str(leakage_path),
        "predictions": str(predictions_path),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="테스트 전용 합성 KCB 데이터·CatBoost 아티팩트 생성")
    parser.add_argument("--rows-per-type", type=int, default=250, help="V1~S3 각 목표 유형의 생성 행 수")
    parser.add_argument("--seed", type=int, default=20260724)
    parser.add_argument("--iterations", type=int, default=250, help="합성 테스트 모델의 CatBoost 최대 반복 수")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--overwrite", action="store_true", help="기존 테스트 합성 CSV를 명시적으로 교체")
    parser.add_argument("--generate-only", action="store_true", help="CSV만 만들고 모델 학습은 생략")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    output = args.output if args.output.is_absolute() else ROOT / args.output
    output = output.resolve()
    dataset_dir = (ROOT / "Dataset").resolve()
    if output.parent != dataset_dir or output.name == "KCB_공식데이터.csv":
        raise ValueError("출력은 Dataset 안의 공식 파일이 아닌 테스트 전용 CSV여야 합니다.")
    if output.exists() and not args.overwrite:
        raise FileExistsError(f"기존 파일을 보호합니다. 다시 만들려면 --overwrite를 사용하세요: {output}")

    config = load_config()
    dictionary = locate_dictionary()
    columns = dictionary_columns(dictionary)
    missing = [column for column in config["columns"]["required"] if column not in columns]
    if missing:
        raise ValueError(f"샘플 46개 컬럼에 현재 필수 계약 컬럼이 없습니다: {missing}")

    frame = generate_frame(args.rows_per_type, args.seed, columns, config)
    generation_validation = validate_generated(frame, config)
    output.parent.mkdir(parents=True, exist_ok=True)
    frame.to_csv(output, index=False, encoding="utf-8-sig")

    sample_candidates = frame[frame["합성목표유형"].eq("V3") & frame["연령대"].eq(25)]
    sample_row = (sample_candidates if not sample_candidates.empty else frame[frame["합성목표유형"].eq("V3")]).iloc[0]
    sample_record = {
        column: int(value) if isinstance(value, (int, np.integer)) or float(value).is_integer() else float(value)
        for column, value in sample_row[config["columns"]["required"]].items()
    }
    sample_path = output.with_name("KCB_테스트_샘플.json")
    write_json(sample_path, sample_record)

    manifest_path = output.with_suffix(".manifest.json")
    manifest: dict[str, Any] = {
        "warning": WARNING,
        "data_kind": "synthetic_test",
        "source_dictionary": str(dictionary.relative_to(ROOT)),
        "source_dictionary_columns": len(columns),
        "seed": args.seed,
        "rows_per_target_type": args.rows_per_type,
        "rows": len(frame),
        "columns": len(frame.columns),
        "csv": str(output.relative_to(ROOT)),
        "csv_sha256": _sha256(output),
        "precise_test_sample": str(sample_path.relative_to(ROOT)),
        "precise_test_sample_target": "V3",
        "generation_validation": generation_validation,
    }
    write_json(manifest_path, manifest)

    if not args.generate_only:
        build_result = build_artifacts(config, output, args.iterations)
        model_path = ROOT / config["project"]["artifacts_dir"] / "model.cbm"
        provenance = {
            **manifest,
            "source_csv": output.name,
            "model_sha256": _sha256(model_path),
            "training_iterations_limit": args.iterations,
            "class_counts": build_result["metrics"]["class_counts"],
            "cv_macro_f1": build_result["metrics"]["best"]["mean_fold_macro_f1"],
            "do_not_use_for_production": True,
        }
        write_json(ROOT / config["project"]["artifacts_dir"] / "model_provenance.json", provenance)
        manifest["artifact_result"] = {
            "model": str(model_path.relative_to(ROOT)),
            "class_counts": provenance["class_counts"],
            "cv_macro_f1": provenance["cv_macro_f1"],
        }
        write_json(manifest_path, manifest)

    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
